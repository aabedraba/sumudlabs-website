from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "Research_Groups_April4th_2026.xlsx"
OUTPUT = ROOT / "lib" / "data" / "research-groups.json"


def slugify(text: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_text.lower()).strip("-")
    return slug


def main() -> None:
    workbook = openpyxl.load_workbook(INPUT, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

    records: list[dict[str, object]] = []
    seen: dict[str, int] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        raw = dict(zip(headers, row))
        tags = [tag.strip() for tag in str(raw["Tags"]).split(";") if tag and str(tag).strip()]
        base_id = slugify(f"{raw['Research Group']}-{raw['University']}-{raw['Country']}")
        seen[base_id] = seen.get(base_id, 0) + 1
        record_id = base_id if seen[base_id] == 1 else f"{base_id}-{seen[base_id]}"

        records.append(
            {
                "id": record_id,
                "name": str(raw["Research Group"]).strip(),
                "website": str(raw["Website"]).strip(),
                "tags": tags,
                "description": str(raw["Description"]).strip(),
                "headOfLab": str(raw["Head of Lab"]).strip(),
                "university": str(raw["University"]).strip(),
                "country": str(raw["Country"]).strip(),
                "city": str(raw["City"]).strip(),
                "source": str(raw["Source"]).strip(),
            }
        )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} research groups to {OUTPUT}")


if __name__ == "__main__":
    main()
